from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from typing import List, Optional
import logging
import uuid
import io
import json
import pandas as pd
from datetime import datetime, timezone
from urllib.parse import quote

from database import db
from models import ProcessTextRequest
from services.matching import find_best_match_improved
from services.processing import evaluate_filter_expression, apply_product_mappings

router = APIRouter()


@router.post("/process-text")
async def process_text_data(request: ProcessTextRequest):
    try:
        # Get store limits
        store = await db.stores.find_one({"id": request.store_id})
        if not store:
            raise HTTPException(status_code=404, detail="Store not found")
        
        limits_dict = {item['product']: item['limit'] for item in store.get('limits', [])}
        
        # Get data either from request or from global stock
        if request.use_global_stock:
            # Load from global stock
            global_stock = await db.global_stock.find_one({}, sort=[("uploaded_at", -1)])
            if not global_stock:
                raise HTTPException(status_code=400, detail="Нет загруженных общих остатков")
            
            store_name = store["name"]
            stock_data = global_stock.get("data", {})
            
            # Get "Электро" stock for warehouse check
            electro_stock = {}
            for product, stores in stock_data.items():
                electro_val = stores.get("Электро", 0)
                electro_stock[product] = electro_val
            
            data_list = []
            removed_by_electro = 0
            for product, stores in stock_data.items():
                # Check Электро warehouse: if (Электро stock - 2) <= 0, skip this product
                electro_val = electro_stock.get(product, 0)
                if electro_val - 2 <= 0:
                    removed_by_electro += 1
                    continue
                
                stock = stores.get(store_name, 0)
                data_list.append({"product": product, "stock": stock})
            
            if removed_by_electro > 0:
                logging.info(f"Removed {removed_by_electro} products - not available on Электро warehouse")
            
            if not data_list:
                raise HTTPException(status_code=400, detail=f"Нет данных для точки '{store_name}' в общих остатках")
            
            data_dict = {
                'Товар': [item["product"] for item in data_list],
                'Остаток': [item["stock"] for item in data_list]
            }
        else:
            # Use provided data
            data_dict = {
                'Товар': [item.product for item in request.data],
                'Остаток': [item.stock for item in request.data]
            }
        
        df = pd.DataFrame(data_dict)
        
        # Process data
        df['Остаток'] = pd.to_numeric(df['Остаток'], errors='coerce').fillna(0)
        df['Товар'] = df['Товар'].astype(str)
        
        # Apply product mappings
        df = await apply_product_mappings(df)
        
        # Save stock history
        history_entries = []
        for _, row in df.iterrows():
            history_entries.append({
                "id": str(uuid.uuid4()),
                "store_id": store["id"],
                "store_name": store["name"],
                "product": row["Товар"],
                "stock": row["Остаток"],
                "recorded_at": datetime.now(timezone.utc).isoformat()
            })
        if history_entries:
            await db.stock_history.insert_many(history_entries)
        
        # Pre-calculate all matches
        logging.info(f"Starting limit matching for {len(df)} products against {len(limits_dict)} limits")
        
        match_cache = {}
        for product in df['Товар'].unique():
            match = find_best_match_improved(product, limits_dict)
            if match and limits_dict[match] > 0:
                match_cache[product] = (match, limits_dict[match])
        
        logging.info(f"Found {len(match_cache)} products with matching limits")
        
        # Filter to only products with limits
        df = df[df['Товар'].isin(match_cache.keys())]
        
        if len(df) == 0:
            raise HTTPException(
                status_code=400, 
                detail="Не найдено товаров с лимитами. Проверьте лимиты и названия товаров."
            )
        
        # Calculate order and limits
        df['Лимиты'] = df['Товар'].apply(lambda x: match_cache[x][1])
        df['Заказ'] = df.apply(lambda row: max(0, match_cache[row['Товар']][1] - row['Остаток']), axis=1)
        
        # Remove zero orders
        df = df[df['Заказ'] > 0]
        
        # Apply custom filters
        if request.filter_expressions:
            for expr in request.filter_expressions:
                if expr.strip():
                    df = df[df.apply(
                        lambda row: evaluate_filter_expression(
                            expr,
                            row['Лимиты'],
                            row['Остаток'],
                            row['Заказ']
                        ),
                        axis=1
                    )]
        
        if len(df) == 0:
            raise HTTPException(
                status_code=400, 
                detail="Не найдено товаров для заказа."
            )
        
        # Append seller request text
        seller_items = []
        if request.seller_request and request.seller_request.strip():
            seller_lines = [line.strip() for line in request.seller_request.strip().split('\n') if line.strip()]
            seller_rows = [{
                'Товар': line,
                'Остаток': 0,
                'Лимиты': 0,
                'Заказ': 0
            } for line in seller_lines]
            
            if seller_rows:
                df = pd.concat([df, pd.DataFrame(seller_rows)], ignore_index=True)
                seller_items = [{
                    "product": row['Товар'],
                    "stock": 0,
                    "order": 0,
                    "limit": 0,
                    "is_seller_request": True
                } for row in seller_rows]
        
        logging.info(f"Final order: {len(df)} items")
        
        # Save order to history
        order_items = []
        for _, row in df.iterrows():
            is_seller = row["Товар"] in [item["product"] for item in seller_items]
            order_items.append({
                "product": row["Товар"],
                "stock": float(row["Остаток"]),
                "order": float(row["Заказ"]),
                "limit": float(row["Лимиты"]),
                "is_seller_request": is_seller
            })
        
        order_history = {
            "id": str(uuid.uuid4()),
            "store_id": store["id"],
            "store_name": store["name"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "items": order_items,
            "seller_request": request.seller_request if request.seller_request else None
        }
        await db.order_history.insert_one(order_history)
        
        # Generate Excel response in the same format as order history download
        store_name = store['name']
        
        # Create simplified format: Store name column + Заказ column
        output_df = pd.DataFrame([
            {store_name: row['Товар'], 'Заказ': int(row['Заказ'])}
            for _, row in df.iterrows()
        ])
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            output_df.to_excel(writer, index=False, sheet_name='Заказ')
            
            # Apply bold font to all cells
            from openpyxl.styles import Font
            workbook = writer.book
            worksheet = writer.sheets['Заказ']
            bold_font = Font(bold=True)
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = bold_font
        
        output.seek(0)
        
        filename = f"{store_name}.xlsx"
        encoded_filename = quote(filename)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Processing error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/process")
async def process_order(
    file: UploadFile = File(...),
    store_id: str = None,
    filter_expressions: str = "[]",
    seller_request: str = ""
):
    """Process order from uploaded Excel file - uses same logic as process_text_data"""
    try:
        filter_list = json.loads(filter_expressions)
        
        store = await db.stores.find_one({"id": store_id})
        if not store:
            raise HTTPException(status_code=404, detail="Store not found")
        
        limits_dict = {item['product']: item['limit'] for item in store.get('limits', [])}
        
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        if 'Товар' not in df.columns or 'Остаток' not in df.columns:
            raise HTTPException(status_code=400, detail="Excel file must contain 'Товар' and 'Остаток' columns")
        
        df['Остаток'] = pd.to_numeric(df['Остаток'], errors='coerce').fillna(0)
        df['Товар'] = df['Товар'].astype(str)
        
        # Apply product mappings (same as process_text_data)
        df = await apply_product_mappings(df)
        
        # Pre-calculate all matches (optimized approach from process_text_data)
        logging.info(f"Starting limit matching for {len(df)} products against {len(limits_dict)} limits")
        
        match_cache = {}
        for product in df['Товар'].unique():
            match = find_best_match_improved(product, limits_dict)
            if match and limits_dict[match] > 0:
                match_cache[product] = (match, limits_dict[match])
        
        logging.info(f"Found {len(match_cache)} products with matching limits")
        
        # Filter to only products with limits
        df = df[df['Товар'].isin(match_cache.keys())]
        
        if len(df) == 0:
            raise HTTPException(
                status_code=400, 
                detail="Не найдено товаров с лимитами. Проверьте лимиты и названия товаров."
            )
        
        # Calculate order and limits
        df['Лимиты'] = df['Товар'].apply(lambda x: match_cache[x][1])
        df['Заказ'] = df.apply(lambda row: max(0, match_cache[row['Товар']][1] - row['Остаток']), axis=1)
        
        # Remove zero orders
        df = df[df['Заказ'] > 0]
        
        # Apply custom filters
        if filter_list:
            for expr in filter_list:
                if expr.strip():
                    df = df[df.apply(
                        lambda row: evaluate_filter_expression(
                            expr,
                            row['Лимиты'],
                            row['Остаток'],
                            row['Заказ']
                        ),
                        axis=1
                    )]
        
        if len(df) == 0:
            raise HTTPException(
                status_code=400, 
                detail="Не найдено товаров для заказа."
            )
        
        # Append seller request text
        seller_items = []
        if seller_request and seller_request.strip():
            seller_lines = [line.strip() for line in seller_request.strip().split('\n') if line.strip()]
            seller_rows = [{
                'Товар': line,
                'Остаток': 0,
                'Лимиты': 0,
                'Заказ': 0
            } for line in seller_lines]
            
            if seller_rows:
                df = pd.concat([df, pd.DataFrame(seller_rows)], ignore_index=True)
                seller_items = [{
                    "product": row['Товар'],
                    "stock": 0,
                    "order": 0,
                    "limit": 0,
                    "is_seller_request": True
                } for row in seller_rows]
        
        logging.info(f"Final order: {len(df)} items")
        
        # Save order to history
        order_items = []
        for _, row in df.iterrows():
            is_seller = row["Товар"] in [item["product"] for item in seller_items]
            order_items.append({
                "product": row["Товар"],
                "stock": float(row["Остаток"]),
                "order": float(row["Заказ"]),
                "limit": float(row["Лимиты"]),
                "is_seller_request": is_seller
            })
        
        order_history = {
            "id": str(uuid.uuid4()),
            "store_id": store["id"],
            "store_name": store["name"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "items": order_items,
            "seller_request": seller_request if seller_request else None
        }
        await db.order_history.insert_one(order_history)
        
        # Generate Excel response in the same format as order history download
        store_name = store['name']
        
        output_df = pd.DataFrame([
            {store_name: row['Товар'], 'Заказ': int(row['Заказ'])}
            for _, row in df.iterrows()
        ])
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            output_df.to_excel(writer, index=False, sheet_name='Заказ')
            
            from openpyxl.styles import Font
            workbook = writer.book
            worksheet = writer.sheets['Заказ']
            bold_font = Font(bold=True)
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = bold_font
        
        output.seek(0)
        
        filename = f"{store_name}.xlsx"
        encoded_filename = quote(filename)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    
    except Exception as e:
        logging.error(f"Processing error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stores/{store_id}/orders")
async def get_store_orders(store_id: str):
    """Get order history for a store"""
    store = await db.stores.find_one({"id": store_id})
    if not store:
        raise HTTPException(status_code=404, detail="Store not found")
    
    orders = await db.order_history.find(
        {"store_id": store_id},
        {"_id": 0, "id": 1, "store_name": 1, "created_at": 1, "items": 1}
    ).sort("created_at", -1).to_list(1000)
    
    for order in orders:
        order["items_count"] = len(order.get("items", []))
    
    return orders


@router.get("/stores/{store_id}/orders/{order_id}")
async def get_order_details(store_id: str, order_id: str):
    """Get details of a specific order"""
    order = await db.order_history.find_one(
        {"store_id": store_id, "id": order_id},
        {"_id": 0}
    )
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


@router.get("/stores/{store_id}/orders/{order_id}/download")
async def download_order(store_id: str, order_id: str):
    """Download order as Excel file"""
    order = await db.order_history.find_one(
        {"store_id": store_id, "id": order_id},
        {"_id": 0}
    )
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    store = await db.stores.find_one({"id": store_id})
    store_name = store["name"] if store else "Заказ"
    
    items = order.get("items", [])
    df = pd.DataFrame([
        {store_name: item.get("product", ""), "Заказ": item.get("order", 0)}
        for item in items
    ])
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Заказ')
        
        workbook = writer.book
        worksheet = writer.sheets['Заказ']
        from openpyxl.styles import Font
        bold_font = Font(bold=True)
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = bold_font
    
    output.seek(0)
    
    filename = f"{store_name}.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )
